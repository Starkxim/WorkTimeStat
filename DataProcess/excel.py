import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from typing import List, Tuple

class CXlAutofit():
    # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
    def get_num_colnum_dict(self):
        '''
        :return: 返回字典：{1:'A', 2:'B', ...... , 52:'AZ'}
        '''
        num_str_dict = {}
        A_Z = [chr(a) for a in range(ord('A'), ord('Z') + 1)]
        AA_AZ = ['A' + chr(a) for a in range(ord('A'), ord('Z') + 1)]
        A_AZ = A_Z + AA_AZ
        for i in A_AZ:
            num_str_dict[A_AZ.index(i) + 1] = i
        return num_str_dict
        
    # 自适应列宽
    def style_excel(self,excel_name:str,sheet_name:str):
        '''
        :param sheet_name:  excel中的sheet名
        :return:
        '''
        wb = load_workbook(excel_name)
        sheet = wb[sheet_name]
        max_column = sheet.max_column
        max_row = sheet.max_row

        # 将每一列，单元格列宽最大的列宽值存到字典里，key:列的序号从1开始(与字典num_str_dic中的key对应)；value:列宽的值
        max_column_dict = {}

        num_str_dict = self.get_num_colnum_dict()

        for i in range(1, max_column + 1):
            for j in range(1, max_row + 1):
                column = 0
                sheet_value = sheet.cell(row=j, column=i).value
                sheet_value_list = [k for k in str(sheet_value)]
                for v in sheet_value_list:
                    # 判定长度，一个数字或一个字母，单元格列宽+=1.1，其它+=2.2（长度可根据需要自行修改，经测试一个字母的列宽长度大概为1）
                    if v.isdigit() == True or v.isalpha() == True:
                        column += 1.1
                    else:
                        column += 2.2
                # 当前单元格列宽与字典中的对比，大于字典中的列宽值则将字典更新。如果字典没有这个key，抛出异常并将值添加到字典中
                try:
                    if column > max_column_dict[i]:
                        max_column_dict[i] = column
                except Exception as e:
                    max_column_dict[i] = column
        # 此时max_column_dict字典中已存有当前sheet的所有列的最大列宽值，直接遍历字典修改列宽
        for key, value in max_column_dict.items():
            sheet.column_dimensions[num_str_dict[key]].width = value + 5 # 5为列宽的额外值，可根据需要自行修改

        wb.save(excel_name)

def read_holiday_data(year: str) -> Tuple[List[str], List[str]]:
    """
    读取节假日和补班日期数据。

    Args:
        year (str): 年份。

    Returns:
        Tuple[List[str], List[str]]: 节假日和补班日期列表。
    """
    public_holidays = pd.read_csv(f"HolidayData/public_holidays_{year}.csv")["date"].tolist()
    makeup_workdays = pd.read_csv(f"HolidayData/makeup_workdays_{year}.csv")["date"].tolist()
    return public_holidays, makeup_workdays


def merge_files(money_file: str, rest_file: str, save_directory: str) -> str:
    """
    合并计薪和调休加班文件，并保存结果。

    Args:
        money_file (str): 计薪加班文件路径。
        rest_file (str): 调休加班文件路径。
        save_directory (str): 保存结果文件的目录。

    Returns:
        str: 保存的结果文件路径。
    """
    try:
        time = datetime.datetime.now().strftime("%Y%M%d%H%M%S")
        output_file = os.path.join(save_directory, f"加班合并结果_{time}.xlsx")

        year = datetime.datetime.now().year.__str__()
        public_holidays, makeup_workdays = read_holiday_data(year)
        public_holidays = pd.to_datetime(public_holidays).date
        makeup_workdays = pd.to_datetime(makeup_workdays).date

        df_money = pd.read_excel(money_file)
        df_rest = pd.read_excel(rest_file)

        df_money = df_money[["申请人", "开始时间", "结束时间", "时长", "项目编号", "当前审批状态"]]
        df_rest = df_rest[["申请人", "开始时间", "结束时间", "加班时长", "项目编号", "当前审批状态"]]

        df_money = df_money[df_money["当前审批状态"] == "已通过"]
        df_rest = df_rest[df_rest["当前审批状态"] == "已通过"]

        df_rest.rename(columns={"加班时长": "时长"}, inplace=True)

        df_money["开始时间"] = pd.to_datetime(df_money["开始时间"], errors="coerce")
        df_rest["开始时间"] = pd.to_datetime(df_rest["开始时间"], errors="coerce")
        df_money["结束时间"] = pd.to_datetime(df_money["结束时间"], errors="coerce")
        df_rest["结束时间"] = pd.to_datetime(df_rest["结束时间"], errors="coerce")

        df_money["类型"] = "计薪"
        df_rest["类型"] = "调休"

        df_combined = pd.concat([df_money, df_rest], ignore_index=True)

        df_combined = df_combined.sort_values(by=["申请人", "类型", "项目编号", "开始时间"], ascending=[True, True, True, True])

        df_combined["时长"] = df_combined["时长"].apply(lambda x: float(str(x).replace("小时", "")) if isinstance(x, str) else x)

        project_money_hours = df_combined[df_combined["类型"] == "计薪"].groupby(["申请人", "项目编号"])["时长"].sum().reset_index()
        project_rest_hours = df_combined[df_combined["类型"] == "调休"].groupby(["申请人", "项目编号"])["时长"].sum().reset_index()

        total_money_hours = df_combined[df_combined["类型"] == "计薪"].groupby("申请人")["时长"].sum().reset_index()
        total_rest_hours = df_combined[df_combined["类型"] == "调休"].groupby("申请人")["时长"].sum().reset_index()

        project_money_hours.rename(columns={"时长": "项目计薪总时长"}, inplace=True)
        project_rest_hours.rename(columns={"时长": "项目调休总时长"}, inplace=True)

        total_money_hours.rename(columns={"时长": "计薪总时长"}, inplace=True)
        total_rest_hours.rename(columns={"时长": "调休总时长"}, inplace=True)

        df_combined = pd.merge(df_combined, project_money_hours, on=["申请人", "项目编号"], how="left")
        df_combined = pd.merge(df_combined, project_rest_hours, on=["申请人", "项目编号"], how="left")

        df_combined = pd.merge(df_combined, total_money_hours, on="申请人", how="left")
        df_combined = pd.merge(df_combined, total_rest_hours, on="申请人", how="left")

        if "类型" in df_combined.columns:
            df_combined["项目调休总时长"] = df_combined.apply(lambda x: 0 if x["类型"] == "计薪" else x["项目调休总时长"], axis=1)
            df_combined["项目计薪总时长"] = df_combined.apply(lambda x: 0 if x["类型"] == "调休" else x["项目计薪总时长"], axis=1)

        df_combined["项目计薪总时长"] = df_combined["项目计薪总时长"].fillna(0)
        df_combined["项目调休总时长"] = df_combined["项目调休总时长"].fillna(0)

        df_combined["计薪总时长"] = df_combined["计薪总时长"].fillna(0)
        df_combined["调休总时长"] = df_combined["调休总时长"].fillna(0)

        df_combined["时长"] = df_combined["时长"].apply(lambda x: f"{x}小时")
        df_combined["项目计薪总时长"] = df_combined["项目计薪总时长"].apply(lambda x: f"{x}小时" if x > 0 else "")
        df_combined["项目调休总时长"] = df_combined["项目调休总时长"].apply(lambda x: f"{x}小时" if x > 0 else "")
        df_combined["计薪总时长"] = df_combined["计薪总时长"].apply(lambda x: f"{x}小时" if x > 0 else "")
        df_combined["调休总时长"] = df_combined["调休总时长"].apply(lambda x: f"{x}小时" if x > 0 else "")

        df_combined["开始时间"] = df_combined["开始时间"].dt.strftime("%Y/%m/%d %H:%M")
        df_combined["结束时间"] = df_combined["结束时间"].dt.strftime("%Y/%m/%d %H:%M")

        df_final = df_combined[["申请人", "开始时间", "结束时间", "时长", "类型", "项目编号", "项目计薪总时长", "项目调休总时长", "调休总时长", "计薪总时长"]]

        df_final.to_excel(output_file, index=False)

        wb = load_workbook(output_file)
        ws = wb.active

        def merge_cells(ws, col: int):
            """
            合并Excel单元格。

            Args:
                ws (Worksheet): Excel工作表对象。
                col (int): 需要合并的列索引。
            """
            max_row = ws.max_row
            merge_start = 2
            for row in range(3, max_row + 2):
                if (
                    ws.cell(row=row, column=col).value != ws.cell(row=row - 1, column=col).value
                    or ws.cell(row=row, column=1).value != ws.cell(row=row - 1, column=1).value
                    or (col in [7, 8] and ws.cell(row=row, column=6).value != ws.cell(row=row - 1, column=6).value)
                ):
                    if merge_start != row - 1:
                        ws.merge_cells(start_row=merge_start, start_column=col, end_row=row - 1, end_column=col)
                        ws.cell(row=merge_start, column=col).alignment = Alignment(vertical="center")
                    merge_start = row

            if merge_start != max_row + 1:
                ws.merge_cells(start_row=merge_start, start_column=col, end_row=max_row + 1, end_column=col)
                ws.cell(row=merge_start, column=col).alignment = Alignment(vertical="center")

        columns_to_merge = [5, 7, 8, 9, 10, 1, 6]
        for col in columns_to_merge:
            merge_cells(ws, col)

        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        ws.freeze_panes = 'A2' # 冻结第一行
        wb.save(output_file)

        # 调用方法 实例化类
        Entity = CXlAutofit()
        # 传入参数：Excel名称，需要设置列宽的Sheet名称
        Entity.style_excel(output_file,'Sheet1')

        return output_file

    except Exception as e:
        raise Exception(f"合并文件时出现错误: {e}")
